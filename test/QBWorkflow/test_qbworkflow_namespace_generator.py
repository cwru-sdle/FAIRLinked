

import pytest
import openpyxl
import os
import tempfile
from unittest.mock import patch, MagicMock
from FAIRLinked.QBWorkflow.utility import NAMESPACE_MAP, HEADER_NAMESPACE, HEADER_BASE_URI
from FAIRLinked.QBWorkflow.namespace_template_generator import generate_namespace_excel  # Adjust import path


class TestGenerateNamespaceExcel:
    """Test suite for generate_namespace_excel function"""

    @pytest.fixture
    def temp_output_file(self):
        """Fixture creating a temporary output file path"""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            file_path = tmp.name
        yield file_path
        # Cleanup after test
        if os.path.exists(file_path):
            os.remove(file_path)

    def test_generate_namespace_excel_basic(self, temp_output_file):
        """Test basic Excel generation with namespace mappings"""
        # Act
        generate_namespace_excel(temp_output_file)
        
        # Assert
        assert os.path.exists(temp_output_file), "Excel file should be created"
        
        # Load and verify the workbook
        wb = openpyxl.load_workbook(temp_output_file)
        ws = wb.active
        
        # Verify basic structure
        assert ws.max_column == 2, "Should have exactly 2 columns"
        assert ws.max_row == len(NAMESPACE_MAP) + 1, f"Should have {len(NAMESPACE_MAP)} data rows + 1 header row"
        
        # Verify headers
        assert ws.cell(row=1, column=1).value == HEADER_NAMESPACE
        assert ws.cell(row=1, column=2).value == HEADER_BASE_URI
        
        # Verify all namespace mappings are present
        expected_namespaces = list(NAMESPACE_MAP.keys())
        expected_uris = list(NAMESPACE_MAP.values())
        
        for row_num, (expected_ns, expected_uri) in enumerate(zip(expected_namespaces, expected_uris), start=2):
            assert ws.cell(row=row_num, column=1).value == expected_ns
            assert ws.cell(row=row_num, column=2).value == expected_uri
        
        wb.close()


    def test_excel_formatting(self, temp_output_file):
        """Test that Excel formatting is correctly applied"""
        # Act
        generate_namespace_excel(temp_output_file)
        
        # Assert
        wb = openpyxl.load_workbook(temp_output_file)
        ws = wb.active
        
        # Test header formatting
        header_cell = ws.cell(row=1, column=1)
        assert header_cell.font.bold is True
        assert header_cell.alignment.horizontal == 'center'
        assert header_cell.alignment.vertical == 'center'
        assert header_cell.alignment.wrap_text is True
        
        # Test header fill colors
        namespace_header = ws.cell(row=1, column=1)
        uri_header = ws.cell(row=1, column=2)
        
        # Check that fills are applied (colors may be stored differently)
        assert namespace_header.fill.fgColor.type == 'rgb'
        assert uri_header.fill.fgColor.type == 'rgb'
        
        # Test borders
        thin_border_style = 'thin'
        test_cell = ws.cell(row=1, column=1)
        assert test_cell.border.left.style == thin_border_style
        assert test_cell.border.right.style == thin_border_style
        assert test_cell.border.top.style == thin_border_style
        assert test_cell.border.bottom.style == thin_border_style
        
        # Test data cell alignment
        data_cell = ws.cell(row=2, column=1)
        assert data_cell.alignment.horizontal == 'center'
        assert data_cell.alignment.vertical == 'center'
        
        # Test URI cell wrap text
        uri_cell = ws.cell(row=2, column=2)
        assert uri_cell.alignment.wrap_text is True
        
        wb.close()


    def test_column_widths_and_row_heights(self, temp_output_file):
        """Test that column widths and row heights are set correctly"""
        # Act
        generate_namespace_excel(temp_output_file)
        
        # Assert
        wb = openpyxl.load_workbook(temp_output_file)
        ws = wb.active
        
        # Test column widths
        assert ws.column_dimensions['A'].width == 15, "Namespace column should be 15 characters wide"
        assert ws.column_dimensions['B'].width == 50, "Base URI column should be 50 characters wide"
        
        # Test row heights for data rows
        for row in range(2, ws.max_row + 1):
            assert ws.row_dimensions[row].height == 40, f"Row {row} should have height 40"
        
        wb.close()

    def test_empty_namespace_map(self, temp_output_file):
        """Test with empty NAMESPACE_MAP"""
        # Mock NAMESPACE_MAP to be empty
        with patch('FAIRLinked.QBWorkflow.namespace_template_generator.NAMESPACE_MAP', {}):
            # Act
            generate_namespace_excel(temp_output_file)
            
            # Assert
            wb = openpyxl.load_workbook(temp_output_file)
            ws = wb.active
            
            # Should only have header row
            assert ws.max_row == 1
            assert ws.max_column == 2
            
            # Headers should still be present
            assert ws.cell(row=1, column=1).value == HEADER_NAMESPACE
            assert ws.cell(row=1, column=2).value == HEADER_BASE_URI
            
            wb.close()

    def test_single_namespace(self, temp_output_file):
        """Test with single namespace in NAMESPACE_MAP"""
        single_map = {'test': 'http://example.com/test'}
        
        with patch('FAIRLinked.QBWorkflow.namespace_template_generator.NAMESPACE_MAP', single_map):
            # Act
            generate_namespace_excel(temp_output_file)
            
            # Assert
            wb = openpyxl.load_workbook(temp_output_file)
            ws = wb.active
            
            assert ws.max_row == 2  # Header + 1 data row
            assert ws.cell(row=2, column=1).value == 'test'
            assert ws.cell(row=2, column=2).value == 'http://example.com/test'
            
            wb.close()

    @patch('builtins.print')
    def test_success_message(self, mock_print, temp_output_file):
        """Test that success message is printed"""
        # Act
        generate_namespace_excel(temp_output_file)
        
        # Assert
        mock_print.assert_called_with(f"Excel file '{temp_output_file}' has been generated with default namespaces.")



    def test_invalid_file_path(self, capsys):
        """Test with invalid file path"""
        invalid_path = "/invalid/path/namespaces.xlsx"
        
        # Act - should not raise exception due to try-except block
        generate_namespace_excel(invalid_path)
        
        # Assert - should print error message instead
        captured = capsys.readouterr()
        assert "An unexpected error occurred while generating the Excel file" in captured.out
        assert "No such file or directory" in captured.out


    @patch('builtins.print')
    def test_exception_handling(self, mock_print):
        """Test exception handling when workbook save fails"""
        # Mock openpyxl to raise an exception
        with patch('openpyxl.Workbook') as mock_workbook:
            mock_wb = MagicMock()
            mock_ws = MagicMock()
            mock_wb.active = mock_ws
            mock_workbook.return_value = mock_wb
            mock_wb.save.side_effect = Exception("Save failed")
            
            # Act
            generate_namespace_excel("/tmp/test.xlsx")
            
            # Assert
            mock_print.assert_called_with("An unexpected error occurred while generating the Excel file: Save failed")

    def test_data_row_alignment(self, temp_output_file):
        """Test that all data rows have proper alignment"""
        # Act
        generate_namespace_excel(temp_output_file)
        
        # Assert
        wb = openpyxl.load_workbook(temp_output_file)
        ws = wb.active
        
        for row in range(2, ws.max_row + 1):
            namespace_cell = ws.cell(row=row, column=1)
            uri_cell = ws.cell(row=row, column=2)
            
            # Both should have center alignment
            assert namespace_cell.alignment.horizontal == 'center'
            assert namespace_cell.alignment.vertical == 'center'
            assert uri_cell.alignment.horizontal == 'center'
            assert uri_cell.alignment.vertical == 'center'
            
            # URI cell should have wrap_text enabled
            assert uri_cell.alignment.wrap_text is True
            
            # Both should have borders
            assert namespace_cell.border.left.style == 'thin'
            assert uri_cell.border.left.style == 'thin'
        
        wb.close()

    def test_header_styling_differences(self, temp_output_file):
        """Test that headers have different styling (different fill colors)"""
        # Act
        generate_namespace_excel(temp_output_file)
        
        # Assert
        wb = openpyxl.load_workbook(temp_output_file)
        ws = wb.active
        
        namespace_header = ws.cell(row=1, column=1)
        uri_header = ws.cell(row=1, column=2)
        
        # Headers should have different fill colors
        assert namespace_header.fill.start_color.index != uri_header.fill.start_color.index
        
        wb.close()

    def test_file_overwrite(self, temp_output_file):
        """Test that function can overwrite existing file"""
        # Create a file first
        with open(temp_output_file, 'w') as f:
            f.write("test content")
        
        # Act - should overwrite the file
        generate_namespace_excel(temp_output_file)
        
        # Assert - file should now be a valid Excel file
        assert os.path.exists(temp_output_file)
        
        # Should be able to load as Excel
        wb = openpyxl.load_workbook(temp_output_file)
        ws = wb.active
        assert ws.max_column == 2
        wb.close()
