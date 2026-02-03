import pytest
import openpyxl
import os
import tempfile
from unittest.mock import patch
from FAIRLinked.QBWorkflow.utility import (
    CATEGORY_COLORS, ALT_LABEL_INSTR,
    UNIT_INSTR, IS_MEASURE_INSTR, EXISTING_URI_INSTR,
    EXPERIMENT_ID_INSTR, NO_TERMS_MSG, TEMPLATE_GENERATED_MSG
)
from FAIRLinked.QBWorkflow.data_template_generator import generate_data_xlsx_template  # Adjust import path


class TestGenerateDataXlsxTemplate:
    """Test suite for generate_data_xlsx_template function"""

    @pytest.fixture
    def sample_children_terms(self):
        """Fixture providing sample children terms with categories"""
        return {
            'mds:tool': ['InstrumentId', 'InstrumentName', 'Manufacturer'],
            'mds:material': ['SampleId', 'SampleType', 'Concentration'],
            'mds:method': ['Protocol', 'Duration', 'Temperature']
        }

    @pytest.fixture
    def empty_children_terms(self):
        """Fixture providing empty children terms"""
        return {}

    @pytest.fixture
    def partial_children_terms(self):
        """Fixture providing children terms with some empty categories"""
        return {
            'mds:tool': ['InstrumentId', 'InstrumentName'],
            'mds:material': [],  # Empty category
            'mds:method': ['Protocol']
        }

    @pytest.fixture
    def temp_output_file(self):
        """Fixture creating a temporary output file path"""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            file_path = tmp.name
        yield file_path
        # Cleanup after test
        if os.path.exists(file_path):
            os.remove(file_path)

    def test_generate_template_with_categories(self, sample_children_terms, temp_output_file):
        """Test generating template with valid categories"""
        # Act
        generate_data_xlsx_template(sample_children_terms, temp_output_file)
        
        # Assert
        assert os.path.exists(temp_output_file), "Excel file should be created"
        
        # Load and verify the workbook
        wb = openpyxl.load_workbook(temp_output_file)
        ws = wb.active
        
        # Verify basic structure
        assert ws.max_row >= 8, "Should have at least 8 rows"
        assert ws.max_column == 12, f"Should have 12 columns (A-L), but has {ws.max_column}"
        
        # Verify category headers
        assert ws.cell(row=1, column=3).value == "Tool"  # mds:tool category (C)
        assert ws.cell(row=1, column=6).value == "Material"  # mds:material category (F)
        assert ws.cell(row=1, column=9).value == "Method"  # mds:method category (I)
        
        # Verify variable names in row 6 (based on function: row_variable_names = 6)
        assert ws.cell(row=6, column=3).value == "InstrumentId"
        assert ws.cell(row=6, column=4).value == "InstrumentName"
        assert ws.cell(row=6, column=5).value == "Manufacturer"
        
        assert ws.cell(row=6, column=6).value == "SampleId"
        assert ws.cell(row=6, column=7).value == "SampleType"
        assert ws.cell(row=6, column=8).value == "Concentration"
        
        assert ws.cell(row=6, column=9).value == "Protocol"
        assert ws.cell(row=6, column=10).value == "Duration"
        assert ws.cell(row=6, column=11).value == "Temperature"
        
        # Verify static instructions (rows 2-5, column A)
        assert ws.cell(row=2, column=1).value == ALT_LABEL_INSTR
        assert ws.cell(row=3, column=1).value == UNIT_INSTR
        assert ws.cell(row=4, column=1).value == IS_MEASURE_INSTR
        assert ws.cell(row=5, column=1).value == EXISTING_URI_INSTR
        
        # Verify ExperimentId and FileName columns
        assert ws.cell(row=1, column=2).value == EXPERIMENT_ID_INSTR
        assert ws.cell(row=6, column=2).value == "ExperimentId"  # Row 6!
        assert ws.cell(row=6, column=12).value == "FileName"  # Column L
        
        # Verify sample data (row_data_start = 7)
        assert ws.cell(row=7, column=2).value == "1"  # ExperimentId
        assert ws.cell(row=7, column=12).value == "Image_1.jpg"  # FileName
        
        wb.close()

    def test_generate_template_empty_categories(self, empty_children_terms, temp_output_file, capsys):
        """Test generating template with empty children terms"""
        # Act
        generate_data_xlsx_template(empty_children_terms, temp_output_file)
        
        # Assert
        captured = capsys.readouterr()
        assert NO_TERMS_MSG in captured.out, "Should print no terms message"
        
        assert os.path.exists(temp_output_file), "Excel file should be created"
        
        # Load and verify the workbook
        wb = openpyxl.load_workbook(temp_output_file)
        ws = wb.active
        
        # Verify Miscellaneous category was created
        assert ws.cell(row=1, column=3).value == "Miscellaneous"
        assert ws.cell(row=6, column=3).value == "var1"
        assert ws.cell(row=6, column=4).value == "var2"
        assert ws.cell(row=6, column=5).value == "var3"
        
        wb.close()

    def test_generate_template_partial_categories(self, partial_children_terms, temp_output_file):
        """Test generating template with some empty categories"""
        # Act
        generate_data_xlsx_template(partial_children_terms, temp_output_file)
        
        # Assert
        assert os.path.exists(temp_output_file), "Excel file should be created"
        
        wb = openpyxl.load_workbook(temp_output_file)
        ws = wb.active
        
        # Should only have mds:tool and mds:method categories (mds:material is empty)
        assert ws.cell(row=1, column=3).value == "Tool"
        assert ws.cell(row=6, column=3).value == "InstrumentId"
        assert ws.cell(row=6, column=4).value == "InstrumentName"
        
        # Method category should be in column 5 (since material is skipped)
        assert ws.cell(row=1, column=5).value == "Method"
        assert ws.cell(row=6, column=5).value == "Protocol"
        
        # Should have 6 columns total: A, B, C, D, E, F
        # A=instructions, B=ExperimentId, C&D=Tool, E=Method, F=FileName
        assert ws.max_column == 6
        
        wb.close()

    def test_template_formatting(self, sample_children_terms, temp_output_file):
        """Test that formatting is correctly applied"""
        # Act
        generate_data_xlsx_template(sample_children_terms, temp_output_file)
        
        # Assert
        wb = openpyxl.load_workbook(temp_output_file)
        ws = wb.active
        
        # Test borders
        thin_border_style = 'thin'
        cell = ws.cell(row=1, column=2)
        assert cell.border.left.style == thin_border_style
        assert cell.border.right.style == thin_border_style
        
        # Test bold font for headers
        assert ws.cell(row=1, column=3).font.bold is True  # Category header
        assert ws.cell(row=6, column=2).font.bold is True  # ExperimentId header - ROW 6!
        
        # Test alignment
        assert ws.cell(row=1, column=2).alignment.horizontal == 'center'
        assert ws.cell(row=1, column=2).alignment.vertical == 'center'
        
        # Test wrap text for instructions
        assert ws.cell(row=2, column=1).alignment.wrap_text is True
        
        # Test row height for instruction rows
        assert ws.row_dimensions[2].height == 60
        assert ws.row_dimensions[3].height == 60
        
        # Test column widths
        assert ws.column_dimensions['A'].width == 80
        assert ws.column_dimensions['B'].width == 20
        
        wb.close()

    def test_category_colors(self, sample_children_terms, temp_output_file):
        """Test that category colors are applied correctly"""
        # Mock CATEGORY_COLORS to ensure consistent testing
        with patch('FAIRLinked.QBWorkflow.utility.CATEGORY_COLORS', {
            'mds:tool': 'FF0000',
            'mds:material': '00FF00',
            'mds:method': '0000FF'
        }):
            # Act
            generate_data_xlsx_template(sample_children_terms, temp_output_file)
            
            # Assert
            wb = openpyxl.load_workbook(temp_output_file)
            ws = wb.active
            
            # Check that colors are applied
            tool_cell = ws.cell(row=1, column=3)
            material_cell = ws.cell(row=1, column=6)
            method_cell = ws.cell(row=1, column=9)
            
            # All should have fill colors
            assert tool_cell.fill.fgColor.type == 'rgb'
            assert material_cell.fill.fgColor.type == 'rgb'
            assert method_cell.fill.fgColor.type == 'rgb'
            
            wb.close()


    def test_merged_cells(self, sample_children_terms, temp_output_file):
        """Test that category headers are properly merged"""
        # Act
        generate_data_xlsx_template(sample_children_terms, temp_output_file)
        
        # Assert
        wb = openpyxl.load_workbook(temp_output_file)
        ws = wb.active
        
        # Check for merged cells - should have exactly 3 merged ranges (one for each category)
        merged_ranges = list(ws.merged_cells.ranges)
        assert len(merged_ranges) == 3, f"Should have exactly 3 merged cells (one per category), but has {len(merged_ranges)}"
        
        # Verify tool category is merged across 3 columns (C, D, E)
        tool_merged = False
        for merged_range in merged_ranges:
            if merged_range.min_row == 1 and merged_range.min_col == 3:
                assert merged_range.max_col == 5  # Columns C-E
                tool_merged = True
                break
        assert tool_merged, "Tool category should be merged across columns C-E"
        
        # Verify material category is merged across 3 columns (F, G, H)
        material_merged = False
        for merged_range in merged_ranges:
            if merged_range.min_row == 1 and merged_range.min_col == 6:
                assert merged_range.max_col == 8  # Columns F-H
                material_merged = True
                break
        assert material_merged, "Material category should be merged across columns F-H"
        
        # Verify method category is merged across 3 columns (I, J, K)
        method_merged = False
        for merged_range in merged_ranges:
            if merged_range.min_row == 1 and merged_range.min_col == 9:
                assert merged_range.max_col == 11  # Columns I-K
                method_merged = True
                break
        assert method_merged, "Method category should be merged across columns I-K"
        
        wb.close()

    @patch('builtins.print')
    def test_template_generated_message(self, mock_print, sample_children_terms, temp_output_file):
        """Test that template generated message is printed"""
        # Act
        generate_data_xlsx_template(sample_children_terms, temp_output_file)
        
        # Assert
        mock_print.assert_called_with(TEMPLATE_GENERATED_MSG.format(temp_output_file))

    def test_file_name_column(self, sample_children_terms, temp_output_file):
        """Test that FileName column is added at the end"""
        # Act
        generate_data_xlsx_template(sample_children_terms, temp_output_file)
        
        # Assert
        wb = openpyxl.load_workbook(temp_output_file)
        ws = wb.active
        
        last_col = ws.max_column
        assert ws.cell(row=1, column=last_col).value == "Data (if any) Supply the name of your data file"
        assert ws.cell(row=6, column=last_col).value == "FileName"  # ROW 6!
        assert ws.cell(row=7, column=last_col).value == "Image_1.jpg"  # ROW 7!
        
        wb.close()

    def test_invalid_output_path(self, sample_children_terms):
        """Test with invalid output path"""
        invalid_path = "/invalid/path/template.xlsx"
        
        # Act & Assert
        with pytest.raises(Exception):
            generate_data_xlsx_template(sample_children_terms, invalid_path)

    def test_single_category_single_term(self, temp_output_file):
        """Test with single category and single term"""
        # Arrange
        children_terms = {'mds:tool': ['InstrumentId']}
        
        # Act
        generate_data_xlsx_template(children_terms, temp_output_file)
        
        # Assert
        wb = openpyxl.load_workbook(temp_output_file)
        ws = wb.active
        
        assert ws.cell(row=1, column=3).value == "Tool"
        assert ws.cell(row=6, column=3).value == "InstrumentId"  # ROW 6!
        
        # Should have 4 columns: A (instructions), B (ExperimentId), C (Tool), D (FileName)
        assert ws.max_column == 4
        
        wb.close()

    def test_category_name_formatting(self, temp_output_file):
        """Test that category names are properly formatted (remove mds:, capitalize)"""
        # Arrange
        children_terms = {
            'mds:some_category': ['Term1'],
            'another:category': ['Term2']  # No mds: prefix
        }
        
        # Act
        generate_data_xlsx_template(children_terms, temp_output_file)
        
        # Assert
        wb = openpyxl.load_workbook(temp_output_file)
        ws = wb.active
        
        assert ws.cell(row=1, column=3).value == "Some_category"  # mds: removed, capitalized
        assert ws.cell(row=1, column=4).value == "Another:category"  # No mds: so unchanged except capitalize
        assert ws.cell(row=6, column=3).value == "Term1"  # ROW 6!
        assert ws.cell(row=6, column=4).value == "Term2"  # ROW 6!
        
        wb.close()
